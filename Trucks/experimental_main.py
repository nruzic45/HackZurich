import pandas as pd
import openpyxl as xl

filename = 'Outbound_ARG.xlsx'

wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# df = pd.read_excel(filename)
#
# column_name = df.columns[4]
# min_we=df[column_name].min()
# max_we=df[column_name].max()
#
# column_name = df.columns[3]
# min_ns=df[column_name].min()
# max_ns=df[column_name].max()


import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QLabel, QMenu, \
    QCheckBox
from PyQt5.QtGui import QPainter, QBrush, QPen
from PyQt5.QtCore import Qt

from generate_graph import inbound_optimize_array, outbound_optimize_array


class Canvas(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setGeometry(0, 0, 400, 300)
        self.setWindowTitle('Canvas')

        self.points = []

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Set up a brush and pen for drawing
        brush = QBrush(Qt.SolidPattern)
        brush.setColor(Qt.blue)
        painter.setBrush(brush)
        pen = QPen(Qt.black, 2, Qt.SolidLine)
        painter.setPen(pen)

        for point in self.points:
            painter.drawEllipse(point[0] - 5, point[1] - 5, 10, 10)

    def clear_canvas(self):
        self.points.clear()
        self.update()


class TextAndCanvasApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # Create labels for text boxes
        label1 = QLabel("City:", self)
        label2 = QLabel("Order:", self)
        label3 = QLabel("Time:", self)

        # Create text boxes arranged horizontally
        self.textbox1 = QLineEdit(self)
        self.textbox2 = QLineEdit(self)
        self.textbox3 = QLineEdit(self)

        self.input_cities = []
        self.values = []
        self.mode = 0

        # Create a layout for text boxes
        textbox_layout = QHBoxLayout()
        textbox_layout.addWidget(label1)
        textbox_layout.addWidget(self.textbox1)
        textbox_layout.addWidget(label2)
        textbox_layout.addWidget(self.textbox2)
        textbox_layout.addWidget(label3)
        textbox_layout.addWidget(self.textbox3)

        # # Create a menu
        # menu = QMenu(self)
        # menu.addAction('Outbound', self.on_option1)
        # menu.addAction('Inbound', self.on_option2)
        #
        # # Create a button for the menu
        # menu_button = QPushButton('Type', self)
        # menu_button.setMenu(menu)

        # Create buttons with events
        self.addButton = QPushButton("Add Order", self)
        self.clearButton = QPushButton("Finish", self)
        self.addButton.clicked.connect(self.add_point)
        self.clearButton.clicked.connect(self.clear_canvas)
        

        # Create canvas
        self.canvas = Canvas()

        # Create a layout for buttons
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.addButton)
        button_layout.addWidget(self.clearButton)

        # Create a label to display the text from the text boxes
        self.text_label = QLabel(self)
        self.text_label.setAlignment(Qt.AlignCenter)

        # Create checkboxes
        self.checkbox1 = QCheckBox("Outbound", self)
        self.checkbox2 = QCheckBox("Inbound", self)

        checkbox_layout = QVBoxLayout()
        checkbox_layout.addWidget(self.checkbox1)
        checkbox_layout.addWidget(self.checkbox2)


        # Create a main layout
        main_layout = QVBoxLayout()
        main_layout.addLayout(checkbox_layout)
        # main_layout.addWidget(menu_button)
        main_layout.addLayout(textbox_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.canvas)
        main_layout.addWidget(self.text_label)

        self.setLayout(main_layout)
        self.setGeometry(100, 100, 400, 400)
        self.setWindowTitle('App for Holcim')




    def on_option1(self):
        print('Option 1 selected')
        self.mode = 0
        print(self.mode)

    def on_option2(self):
        print('Option 2 selected')
        self.mode = 1

    def add_point(self):
        city = self.textbox1.text()
        value = self.textbox2.text()

        self.input_cities.append(city)
        self.values.append(value)

        x = self.canvas.width() // 2
        y = self.canvas.height() // 2

        flag = False

        for i in range(1, 15200):
            if (ws1.cell(row=i, column=3).internal_value == city):
                x = ws1.cell(row=i, column=4).internal_value
                y = ws1.cell(row=i, column=5).internal_value

                x = self.canvas.width() // 2 + x
                y = self.canvas.height() // 2 + y

                flag = True
                break

        if not flag:
            self.text_label.setText('Wrong city name. Please try again.')
            self.textbox1.clear()
            self.textbox2.clear()
            self.textbox3.clear()
            return

        self.canvas.points.append((x, y))
        self.canvas.update()

        # Read and display the text from the text boxes
        text1 = self.textbox1.text()
        text2 = self.textbox2.text()
        text3 = self.textbox3.text()
        

        self.text_label.setText(f'City: {text1}\nOrder: {text2}\nTime: {text3}')

        self.textbox1.clear()
        self.textbox2.clear()
        self.textbox3.clear()

    def clear_canvas(self):
        
        in_arr = self.conv_2_tuple()
        res1 = inbound_optimize_array(in_arr)
        res2 = outbound_optimize_array(in_arr)

        to_front_Inbound = "Inbound sources: " + self.array_to_string(res1[0]) + "Percantage: " + self.array_to_string(res1[1])
        to_front_Outbound = "Outbound factories: " + self.array_to_string(res2[0])

        print(res1)
        print(res2)

        self.text_label.setText(to_front_Inbound)
        self.text_label.setText(to_front_Outbound)
        self.canvas.update()

        self.canvas.clear_canvas()
        self.text_label.clear()

    def conv_2_tuple(self):
        
        input_arr = []
        for i in range(0,len(self.input_cities)):
            tupl = (self.input_cities[i], self.values[i])
            input_arr.append(tupl)

        return input_arr

    def array_to_string(self,arr):
    
        result = ''.join(map(str, arr))
        return result


def main():
    app = QApplication(sys.argv)
    window = TextAndCanvasApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()

